"""
backtest.py  —  Walk-Forward Backtester
========================================
Runs all 6 Albert-Ray strategies on 3 years of full-market history.
No look-ahead bias: each signal is detected on data available up to that day,
entry is simulated at the NEXT bar's open.

Output: Backtest_Results.xlsx  (saved one folder above trading_bot)
  - Summary        per-strategy win rate, profit factor, avg R, max drawdown
  - All Trades     every signal with full detail and outcome
  - Equity Curve   cumulative R over time (all strategies combined)
  - Outcome Breakdown  SL / TP1 / TP2 / EMA trail / max-hold counts

Usage:
  python backtest.py
  (or double-click RUN BACKTEST.bat)

Runtime estimate: 30–90 minutes depending on universe size.
A cache file (backtest_cache.pkl) is saved so you can restart without
re-fetching data you already downloaded.
"""
import logging
import os
import pickle
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import pytz
import requests

# ─── bootstrap path so we can import bot modules ─────────────────────────────
BOT_DIR = Path(__file__).parent
if str(BOT_DIR) not in sys.path:
    sys.path.insert(0, str(BOT_DIR))

from dotenv import load_dotenv
load_dotenv(BOT_DIR / ".env")

from config import Config
from indicators import add_emas
from pattern_detector import scan_symbol
import decision_logger as _dlog

# ─── settings ────────────────────────────────────────────────────────────────
YEARS          = 3          # years of history to fetch
LOOKBACK_DAYS  = YEARS * 365 + 30   # extra buffer for API
MIN_BARS       = 120        # skip symbol if fewer bars available
MAX_HOLD_DAYS  = 30         # force-close after this many days
COOLDOWN_DAYS  = 15         # don't re-signal same stock+strategy within N days
CACHE_FILE     = BOT_DIR / "backtest_cache.pkl"
OUT_DIR        = BOT_DIR.parent           # one level above trading_bot
OUT_PATH       = OUT_DIR / "Backtest_Results.xlsx"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("backtest")
ET  = pytz.timezone("America/New_York")


# ─── Data helpers ─────────────────────────────────────────────────────────────

def _headers() -> dict:
    return {
        "APCA-API-KEY-ID":     Config.ALPACA_API_KEY,
        "APCA-API-SECRET-KEY": Config.ALPACA_SECRET_KEY,
    }


def _fetch_bars(symbol: str) -> pd.DataFrame | None:
    """Fetch up to LOOKBACK_DAYS of daily bars for one symbol."""
    end   = datetime.now(ET).strftime("%Y-%m-%d")
    start = (datetime.now(ET) - timedelta(days=LOOKBACK_DAYS)).strftime("%Y-%m-%d")
    url   = f"https://data.alpaca.markets/v2/stocks/{symbol}/bars"
    params = {
        "timeframe": "1Day",
        "start":     start,
        "end":       end,
        "limit":     1500,
        "feed":      "iex",
    }
    try:
        resp = requests.get(url, headers=_headers(), params=params, timeout=15)
        if resp.status_code == 422:          # symbol not supported on IEX
            return None
        resp.raise_for_status()
        data = resp.json().get("bars", [])
        if not data:
            return None
        df = pd.DataFrame(data)
        df["t"] = pd.to_datetime(df["t"])
        df = df.rename(columns={"t":"date","o":"open","h":"high",
                                 "l":"low","c":"close","v":"volume"})
        df = df.set_index("date").sort_index()
        df = df[["open","high","low","close","volume"]].astype(float)
        df["volume"] = df["volume"].astype(int)
        return df
    except Exception as e:
        log.debug(f"  [fetch] {symbol}: {e}")
        return None


def _get_universe() -> list[str]:
    """Pull full market universe using same logic as the live bot."""
    log.info("Building universe (full market scan)…")
    url    = f"{Config.ALPACA_BASE_URL}/v2/assets"
    params = {"status": "active", "asset_class": "us_equity", "tradable": True}
    try:
        resp = requests.get(url, headers=_headers(), params=params, timeout=30)
        resp.raise_for_status()
        assets = resp.json()
    except Exception as e:
        log.error(f"Asset fetch failed: {e}")
        return list(Config.BASE_WATCHLIST)

    symbols = [
        a["symbol"] for a in assets
        if a.get("tradable")
        and 1 <= len(a["symbol"]) <= 5
        and a["symbol"].isalpha()
    ]
    log.info(f"  {len(symbols)} tradeable symbols → snapshot pre-filter…")

    # Batch snapshot pre-filter
    BATCH    = 100
    snap_url = "https://data.alpaca.markets/v2/stocks/snapshots"
    qualified: list[str] = []

    for i in range(0, len(symbols), BATCH):
        batch = symbols[i: i + BATCH]
        try:
            r = requests.get(snap_url, headers=_headers(),
                             params={"symbols": ",".join(batch), "feed": "iex"},
                             timeout=15)
            r.raise_for_status()
            for sym, snap in r.json().items():
                try:
                    day  = snap.get("dailyBar") or snap.get("prevDailyBar") or {}
                    price  = float(snap.get("latestTrade", {}).get("p", 0) or day.get("c", 0))
                    volume = float(day.get("v", 0))
                    h, lo, c = float(day.get("h",0)), float(day.get("l",0)), float(day.get("c",1))
                    adr = (h - lo) / c * 100 if c else 0
                    if price >= Config.MIN_PRICE and volume >= Config.MIN_AVG_VOLUME and adr >= Config.MIN_ADR_PCT:
                        qualified.append(sym)
                except Exception:
                    continue
        except Exception as e:
            log.debug(f"  snapshot batch {i}: {e}")
        time.sleep(0.05)

    # Always include core watchlist
    for sym in Config.BASE_WATCHLIST:
        if sym not in qualified:
            qualified.append(sym)

    log.info(f"  Universe: {len(qualified)} symbols after pre-filter")
    return qualified


# ─── Data cache ───────────────────────────────────────────────────────────────

def load_cache() -> dict[str, pd.DataFrame]:
    if CACHE_FILE.exists():
        try:
            with open(CACHE_FILE, "rb") as f:
                cache = pickle.load(f)
            log.info(f"Loaded cache: {len(cache)} symbols already fetched")
            return cache
        except Exception:
            pass
    return {}


def save_cache(cache: dict):
    try:
        with open(CACHE_FILE, "wb") as f:
            pickle.dump(cache, f)
    except Exception as e:
        log.warning(f"Cache save failed: {e}")


# ─── Trade simulation ─────────────────────────────────────────────────────────

def simulate_trade(df: pd.DataFrame, signal_idx: int,
                   entry_plan: float, stop: float, tp1: float, tp2: float,
                   direction: str) -> dict:
    """
    Simulate from bar AFTER signal (entry at next open).
    Tracks: SL, TP1 (move stop to BE), TP2, 8-EMA trail break, max hold.
    Returns outcome dict.
    """
    is_long  = direction == "long"
    start    = signal_idx + 1          # entry bar
    if start >= len(df):
        return {"outcome": "no_data", "pnl_r": 0.0, "days_held": 0, "tp1_hit": False}

    actual_entry = df.iloc[start]["open"]
    risk = abs(actual_entry - stop)
    if risk < 0.01:
        risk = abs(entry_plan - stop) or 0.01

    # After TP1 hit, the effective stop moves to entry (breakeven)
    tp1_hit       = False
    current_stop  = stop    # starts at original stop, moves to BE after TP1

    for i in range(start, min(start + MAX_HOLD_DAYS, len(df))):
        bar        = df.iloc[i]
        bar_open   = bar["open"]
        bar_high   = bar["high"]
        bar_low    = bar["low"]
        bar_close  = bar["close"]
        ema8       = bar.get("ema8", None)
        days_held  = i - start + 1

        if is_long:
            # ── SL check (worst case: low hits stop) ──────────────────
            # If we gap below stop on open, exit at open
            if bar_open <= current_stop:
                exit_p = bar_open
                pnl_r  = (exit_p - actual_entry) / risk
                return {"outcome": "sl_hit", "pnl_r": round(pnl_r, 2),
                        "days_held": days_held, "tp1_hit": tp1_hit,
                        "exit_price": round(exit_p, 2)}
            if bar_low <= current_stop:
                exit_p = current_stop
                pnl_r  = (exit_p - actual_entry) / risk
                outcome = "be_stop" if tp1_hit else "sl_hit"
                return {"outcome": outcome, "pnl_r": round(pnl_r, 2),
                        "days_held": days_held, "tp1_hit": tp1_hit,
                        "exit_price": round(exit_p, 2)}

            # ── TP2 check ─────────────────────────────────────────────
            if tp1_hit and bar_high >= tp2:
                pnl_r = (tp2 - actual_entry) / risk
                return {"outcome": "tp2_hit", "pnl_r": round(pnl_r, 2),
                        "days_held": days_held, "tp1_hit": True,
                        "exit_price": round(tp2, 2)}

            # ── TP1 check ─────────────────────────────────────────────
            if not tp1_hit and bar_high >= tp1:
                tp1_hit      = True
                current_stop = actual_entry     # move stop to BE

            # ── 8-EMA trail (only after TP1 and once past bar 2) ─────
            if tp1_hit and ema8 is not None and days_held >= 2:
                if bar_close < ema8:
                    pnl_r = (bar_close - actual_entry) / risk
                    return {"outcome": "ema_trail", "pnl_r": round(pnl_r, 2),
                            "days_held": days_held, "tp1_hit": True,
                            "exit_price": round(bar_close, 2)}

        else:  # SHORT
            if bar_open >= current_stop:
                exit_p = bar_open
                pnl_r  = (actual_entry - exit_p) / risk
                return {"outcome": "sl_hit", "pnl_r": round(pnl_r, 2),
                        "days_held": days_held, "tp1_hit": tp1_hit,
                        "exit_price": round(exit_p, 2)}
            if bar_high >= current_stop:
                exit_p = current_stop
                pnl_r  = (actual_entry - exit_p) / risk
                outcome = "be_stop" if tp1_hit else "sl_hit"
                return {"outcome": outcome, "pnl_r": round(pnl_r, 2),
                        "days_held": days_held, "tp1_hit": tp1_hit,
                        "exit_price": round(exit_p, 2)}

            if tp1_hit and bar_low <= tp2:
                pnl_r = (actual_entry - tp2) / risk
                return {"outcome": "tp2_hit", "pnl_r": round(pnl_r, 2),
                        "days_held": days_held, "tp1_hit": True,
                        "exit_price": round(tp2, 2)}

            if not tp1_hit and bar_low <= tp1:
                tp1_hit      = True
                current_stop = actual_entry

            if tp1_hit and ema8 is not None and days_held >= 2:
                if bar_close > ema8:
                    pnl_r = (actual_entry - bar_close) / risk
                    return {"outcome": "ema_trail", "pnl_r": round(pnl_r, 2),
                            "days_held": days_held, "tp1_hit": True,
                            "exit_price": round(bar_close, 2)}

    # Max hold reached: exit at last bar's close
    last_idx  = min(start + MAX_HOLD_DAYS - 1, len(df) - 1)
    last_close = df.iloc[last_idx]["close"]
    pnl_r = ((last_close - actual_entry) / risk if is_long
              else (actual_entry - last_close) / risk)
    return {"outcome": "max_hold", "pnl_r": round(pnl_r, 2),
            "days_held": MAX_HOLD_DAYS, "tp1_hit": tp1_hit,
            "exit_price": round(last_close, 2)}


# ─── Per-symbol walk-forward ───────────────────────────────────────────────────

def backtest_symbol(symbol: str, df: pd.DataFrame) -> list[dict]:
    """Walk forward through df, run all 6 detectors, simulate each signal."""
    df = add_emas(df.copy())
    trades: list[dict] = []
    # cooldown: strategy -> last bar index where we fired
    cooldowns: dict[str, int] = {}

    for i in range(MIN_BARS, len(df) - 1):
        window = df.iloc[: i + 1].copy()   # only past data
        date   = df.index[i]

        try:
            signals = scan_symbol(symbol, window, df_intraday=None)
        except Exception:
            continue

        for sig in signals:
            strat = sig.get("strategy", "Unknown")
            last  = cooldowns.get(strat, -9999)
            if i - last < COOLDOWN_DAYS:
                continue              # still in cooldown for this strategy
            cooldowns[strat] = i

            result = simulate_trade(
                df, i,
                sig["entry"], sig["stop"], sig["tp1"], sig["tp2"],
                sig.get("signal_type", "long"),
            )

            trade_row = {
                "symbol":      symbol,
                "strategy":    strat,
                "direction":   sig.get("signal_type", "long").upper(),
                "signal_date": date.strftime("%Y-%m-%d"),
                "entry_plan":  round(sig["entry"], 2),
                "stop":        round(sig["stop"],  2),
                "tp1":         round(sig["tp1"],   2),
                "tp2":         round(sig["tp2"],   2),
                "vol_ratio":   round(sig.get("volume_ratio", 0), 2),
                "adr_pct":     round(sig.get("adr_pct", 0),    2),
                "notes":       sig.get("notes", ""),
                **result,
            }
            trades.append(trade_row)

            # Log trade outcome to backtest_journal.txt
            dlog = _dlog.get_logger()
            if dlog:
                dlog.trade_outcome(
                    symbol       = symbol,
                    strategy     = strat,
                    signal_date  = date.strftime("%Y-%m-%d"),
                    entry        = round(sig["entry"], 2),
                    exit_price   = result.get("exit_price", sig["entry"]),
                    pnl_r        = result.get("pnl_r", 0),
                    outcome      = result.get("outcome", "unknown"),
                    days_held    = result.get("days_held", 0),
                )

    return trades


# ─── Main ─────────────────────────────────────────────────────────────────────

def run_backtest() -> list[dict]:
    # Enable decision logger in backtest mode so every signal gets explained
    _dlog.set_logger(_dlog.DecisionLogger(mode="backtest"))
    _dlog._active.mode = "backtest"   # suppress rejections — too noisy at scale
    import os; os.environ.setdefault("DECISION_LOG_LEVEL", "signals_only")

    log.info("=" * 65)
    log.info("  WALK-FORWARD BACKTEST  |  3 Years  |  Full Market")
    log.info("=" * 65)

    symbols = _get_universe()
    cache   = load_cache()
    all_trades: list[dict] = []
    total = len(symbols)

    for idx, sym in enumerate(symbols, 1):
        prefix = f"[{idx:>4}/{total}] {sym:<6}"

        # Use cached bars if available
        if sym in cache:
            df = cache[sym]
            log.info(f"{prefix}  (cached, {len(df)} bars)")
        else:
            df = _fetch_bars(sym)
            if df is not None:
                cache[sym] = df
            # Save cache every 50 symbols in case of crash
            if idx % 50 == 0:
                save_cache(cache)
            # Light throttle: ~4 req/sec to avoid rate-limit
            time.sleep(0.25)

        if df is None or len(df) < MIN_BARS:
            log.info(f"{prefix}  skip (insufficient data)")
            continue

        trades = backtest_symbol(sym, df)
        all_trades.extend(trades)

        if trades:
            wins = sum(1 for t in trades if t["pnl_r"] > 0)
            log.info(f"{prefix}  {len(trades)} signals | {wins} wins | "
                     f"Total R: {sum(t['pnl_r'] for t in trades):+.2f}")
        else:
            log.info(f"{prefix}  0 signals")

    save_cache(cache)
    log.info(f"\nDone. Total signals across all symbols: {len(all_trades)}")
    return all_trades


# ─── Excel report builder ─────────────────────────────────────────────────────

def _col_widths(ws):
    """Auto-size all columns in an openpyxl worksheet."""
    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(width + 3, 45)


def build_report(trades: list[dict]):
    if not trades:
        log.error("No trades to report — nothing was saved.")
        return

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        log.error("openpyxl not installed. Run: pip install openpyxl")
        return

    df = pd.DataFrame(trades)

    # ── 1. Summary by strategy ────────────────────────────────────────────────
    def _summarise(grp: pd.DataFrame) -> pd.Series:
        n        = len(grp)
        winners  = grp[grp["pnl_r"] > 0]
        losers   = grp[grp["pnl_r"] <= 0]
        win_r    = len(winners) / n * 100
        tp1_pct  = grp["tp1_hit"].mean() * 100 if "tp1_hit" in grp else 0
        avg_win  = winners["pnl_r"].mean()  if len(winners) else 0
        avg_loss = losers["pnl_r"].mean()   if len(losers) else 0
        total_r  = grp["pnl_r"].sum()
        gross_p  = winners["pnl_r"].sum()
        gross_l  = abs(losers["pnl_r"].sum())
        pf       = gross_p / gross_l if gross_l > 0 else 999.0
        # Running drawdown (by signal date order)
        cumR     = grp.sort_values("signal_date")["pnl_r"].cumsum()
        peak     = cumR.cummax()
        dd       = (cumR - peak).min()
        return pd.Series({
            "Signals":        n,
            "Win Rate %":     round(win_r, 1),
            "TP1 Hit %":      round(tp1_pct, 1),
            "Avg Win (R)":    round(avg_win,  2),
            "Avg Loss (R)":   round(avg_loss, 2),
            "Profit Factor":  round(pf,        2),
            "Total R":        round(total_r,   2),
            "Max Drawdown R": round(dd,        2),
            "Avg Days Held":  round(grp["days_held"].mean(), 1),
        })

    summary = df.groupby("strategy").apply(_summarise).reset_index()
    # Overall row
    ov = _summarise(df)
    ov["strategy"] = "★ OVERALL"
    summary = pd.concat([summary, pd.DataFrame([ov])], ignore_index=True)
    summary = summary.rename(columns={"strategy": "Strategy"})

    # ── 2. Equity curve (all strategies combined, chronological) ─────────────
    df_sorted = df.sort_values("signal_date").copy()
    df_sorted["cumulative_R"] = df_sorted["pnl_r"].cumsum()
    equity = df_sorted[["signal_date","symbol","strategy","direction",
                         "pnl_r","cumulative_R","outcome"]].copy()

    # ── 3. Outcome breakdown ──────────────────────────────────────────────────
    oc = df["outcome"].value_counts().reset_index()
    oc.columns = ["Outcome", "Count"]
    oc["Pct %"] = (oc["Count"] / len(df) * 100).round(1)

    # ── Write workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT  = Font(color="FFFFFF", bold=True)
    WIN_FILL     = PatternFill("solid", fgColor="C6EFCE")
    LOSS_FILL    = PatternFill("solid", fgColor="FFC7CE")
    OVERALL_FILL = PatternFill("solid", fgColor="BDD7EE")

    def _write_sheet(wb, name: str, data: pd.DataFrame,
                     highlight_col: str = None):
        ws = wb.create_sheet(name)
        rows = list(dataframe_to_rows(data, index=False, header=True))
        for r_idx, row in enumerate(rows, 1):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(r_idx, c_idx, val)
                if r_idx == 1:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal="center")
                elif highlight_col and r_idx > 1:
                    col_names = list(data.columns)
                    if highlight_col in col_names:
                        hc = col_names.index(highlight_col) + 1
                        try:
                            v = float(ws.cell(r_idx, hc).value or 0)
                            if "OVERALL" in str(ws.cell(r_idx, 1).value or ""):
                                ws.cell(r_idx, c_idx).fill = OVERALL_FILL
                            elif v > 0:
                                cell.fill = WIN_FILL
                            elif v < 0:
                                cell.fill = LOSS_FILL
                        except Exception:
                            pass
        _col_widths(ws)
        return ws

    _write_sheet(wb, "Summary",           summary, "Total R")
    _write_sheet(wb, "All Trades",        df,       "pnl_r")
    _write_sheet(wb, "Equity Curve",      equity,   "cumulative_R")
    _write_sheet(wb, "Outcome Breakdown", oc)

    wb.save(OUT_PATH)
    log.info(f"\n{'='*65}")
    log.info(f"  Report saved → {OUT_PATH}")
    log.info(f"{'='*65}")

    # Print quick summary to console
    log.info("\nQUICK SUMMARY:")
    ov_row = summary[summary["Strategy"] == "★ OVERALL"].iloc[0]
    log.info(f"  Total signals  : {int(ov_row['Signals'])}")
    log.info(f"  Win rate       : {ov_row['Win Rate %']}%")
    log.info(f"  Profit factor  : {ov_row['Profit Factor']}")
    log.info(f"  Total R gained : {ov_row['Total R']}")
    log.info(f"  Max drawdown   : {ov_row['Max Drawdown R']} R")
    log.info("")
    log.info("BY STRATEGY:")
    for _, row in summary[summary["Strategy"] != "★ OVERALL"].iterrows():
        log.info(f"  {row['Strategy']:<35}  {row['Signals']:>4} signals  "
                 f"{row['Win Rate %']:>5.1f}% win  PF {row['Profit Factor']:>5.2f}  "
                 f"Total R {row['Total R']:>+7.2f}")


# ─── Entry point ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    start_time = time.time()
    trades     = run_backtest()
    build_report(trades)
    elapsed = (time.time() - start_time) / 60
    log.info(f"\nTotal runtime: {elapsed:.1f} minutes")
    input("\nPress Enter to close…")
